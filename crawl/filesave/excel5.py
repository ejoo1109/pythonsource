from openpyxl import Workbook
import openpyxl.utils.cell as utils
from openpyxl.styles import Alignment, Font  # Alignment: 정렬

# 셀 서식 지정하여 엑셀파일 만들기

# 객체 생성-기본 시트 자동생성됨
excel_file = Workbook()
# 기본 시트 이름 확인
print(excel_file.sheetnames)
# 기본 시트 삭제
excel_file.remove(excel_file["Sheet"])

# 첫번째 시트 생성 : index= 위치, title=시트이름
sheet1 = excel_file.create_sheet(index=0, title="Column")
# 두번째 시트 생성
sheet2 = excel_file.create_sheet(index=1, title="매출표")
# 생성된 시트 확인
print(excel_file.sheetnames)

# 첫번째 시트 3행 6열(col : 열, row : 행)
for col in sheet1.iter_cols(min_col=1, max_col=6, min_row=1, max_row=3):
    for each_cell in col:
        # 값 지정 get_column_letter: 현재컬럼명
        each_cell.value = utils.get_column_letter(each_cell.column)
        # 정렬 기준 설정 horizontal: 수평정렬, vertical: 수직정렬
        each_cell.alignment = Alignment(horizontal="right", vertical="center")
        # 폰트 설정
        each_cell.font = Font(
            bold=True, name="Arial", size=12, underline="single", color="1bb638"
        )


# 두번째 시트 6행 3열
for col in sheet2.iter_cols(min_col=1, max_col=3, min_row=1, max_row=6):
    for each_cell in col:
        # 값 지정
        each_cell.value = each_cell.row
        # 정렬 기준 설정
        each_cell.alignment = Alignment(horizontal="center", vertical="center")
        # 폰트 설정
        each_cell.font = Font(italic=True, name="Consoras", size=10, color="ff0000")


excel_file.save("./crawl/data/test2.xlsx")  # 저장

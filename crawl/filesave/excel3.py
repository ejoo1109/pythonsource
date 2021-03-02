from openpyxl import Workbook

# 엑셀 파일 저장하기

# 객체 생성-기본 시트 자동생성됨
excel_file = Workbook()
# 기본 시트 이름 확인
print(excel_file.sheetnames)
# 기본 시트 삭제
excel_file.remove(excel_file["Sheet"])

# 첫번째 시트 생성 : index= 위치, title=시트이름
sheet1 = excel_file.create_sheet(index=0, title="Column")
# 두번째 시트 생성
sheet2 = excel_file.create_sheet(index=1, title="매출표")
# 생성된 시트 확인
print(excel_file.sheetnames)

excel_file.save("./crawl/data/test2.xlsx")  # 저장

# gmarket best [컴퓨터/전자] 카테고리 제품명 출력
import requests
from bs4 import BeautifulSoup
import openpyxl

# 엑셀 파일로 저장 준비
gmarket_best = openpyxl.Workbook()  # openpyxl의 Workbook()클래스

# 기본 시트 활성화
sheet1 = gmarket_best.active

# 시트명 지정
sheet1.title = "베스트100"

# 열 너비 조정
sheet1.column_dimensions["B"].width = 30
sheet1.column_dimensions["C"].width = 30
sheet1.column_dimensions["D"].width = 30
sheet1.column_dimensions["E"].width = 30

# 헤더명 지정
sheet1.append(["번호", "회사명", "제품명", "상세정보 url", "가격"])

url = "http://corners.gmarket.co.kr/Bestsellers?viewType=G&groupCode=G06"

with requests.Session() as s:
    r = s.get(url)

    soup = BeautifulSoup(r.text, "html.parser")

    # gBestWrap > div > div:nth-child(5) > div:nth-child(3) > ul >li
    best_list = soup.select(
        "#gBestWrap > div > div:nth-child(5) > div:nth-child(3) > ul > li"
    )
    print(best_list)

    for idx, item in enumerate(best_list, 1):  # enumerate 순번부여
        # print("{}".format(item.find('a').string))
        title = item.find("img")["alt"]
        price = item.select_one(".s-price > strong").text

        # 2차 크롤링 (상세페이지에서 판매자 정보)
        product_url = item.find("a")["href"]  # a태그가 가지고 있는 주소값 출력하기
        product_detail = s.get(product_url)  # 상세페이지
        product_soup = BeautifulSoup(product_detail.text, "html.parser")  # 객체 생성
        company = product_soup.select_one("p.shoptit a").string

        # container > div.item-topinfowrap > div.item-topinfo.item-topinfo--additional > div.item-topinfo_headline > p > span.text__seller

        print("{}. {} {} {} {}".format(idx, title, price, product_url, company))

        # 시트에 추가
        sheet1.append([idx, company, title, product_url, price])

        # 주소가 들어 있는 셀에 하이퍼링크 걸어주기
        sheet1.cell(row=idx + 1, column=4).hyperlink = product_url
        # (row=idx + 1, column=4) 행번호는 한번 for 돌때마다 증가해야한다.

    # 셀 서식 지정 - A1셀
    cell_A1 = sheet1["A1"]
    # 중앙정렬
    cell_A1.alignment = openpyxl.styles.Alignment(horizontal="center")
    # 글자색상변경
    cell_A1.font = openpyxl.styles.Font(color="01579B")

    # 셀 서식 지정 - B1셀
    cell_B1 = sheet1["B1"]
    # 중앙정렬
    cell_B1.alignment = openpyxl.styles.Alignment(horizontal="center")
    # 글자색상변경
    cell_B1.font = openpyxl.styles.Font(color="01579B")

    # 셀 서식 지정 - C1셀
    cell_C1 = sheet1["C1"]
    # 중앙정렬
    cell_C1.alignment = openpyxl.styles.Alignment(horizontal="center")
    # 글자색상변경
    cell_C1.font = openpyxl.styles.Font(color="01579B")

    # 셀 서식 지정 - D1셀
    cell_D1 = sheet1["D1"]
    # 중앙정렬
    cell_D1.alignment = openpyxl.styles.Alignment(horizontal="center")
    # 글자색상변경
    cell_D1.font = openpyxl.styles.Font(color="01579B")

    # 셀 서식 지정 - E1셀
    cell_E1 = sheet1["E1"]
    # 중앙정렬
    cell_E1.alignment = openpyxl.styles.Alignment(horizontal="center")
    # 글자색상변경
    cell_E1.font = openpyxl.styles.Font(color="01579B")

    gmarket_best.save("./crawl/data/gmarket_best.xlsx")
    gmarket_best.close()